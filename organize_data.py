import os
import argparse
import openpyxl
import shutil

# Function to extract name, category, weeks, and segmentation filename for a given sample name
def get_name_category_week_segmentation(row):
    name = row[0].value   # Name
    category = row[3].value  # Category
    weeks = row[4].value  # Weeks
    segmentation_fname = row[2].value  # Cine information (Segmentation filename)
    return name, category, weeks, segmentation_fname

# Function to extract PV filename from the PV data column (2nd column)
def get_PV_fname(PV_data):
    # Find the last backslash in the path
    last_backslash_index = PV_data.rfind('\\')
    
    # Extract the file name after the last backslash and remove the last 7 characters
    if last_backslash_index != -1:
        PV_fname = PV_data[last_backslash_index + 1:-7]
        return PV_fname
    else:
        return None

# Function to get the output directory path based on name, category, and weeks
def get_output_directory(directory_path, sample_name, category, weeks):
    # x is the weeks value minus the last character
    x = weeks[:-1]
    
    # s is the sample name with the second to last character replaced by '_'
    s = sample_name[:-2] + "_" + sample_name[-1]
    
    # Check the category to determine the output directory
    if category == "Sham":
        # Build the output directory for Sham category
        outdir = os.path.join(directory_path, f'SHAM/{x}weeks/{s}')
    else:
        # Convert the category value to double and multiply by 100
        c = float(category.replace(',', '.')) * 100
        # Build the output directory for non-Sham category
        outdir = os.path.join(directory_path, f'AS/{x}weeks/{c:.0f}/{s}')
    
    return outdir

# Function to copy segmentation file to the created directory, adding .mat extension to the filename
def copy_segmentation_file(segmentation_fname, segmentation_directory, dest_directory):
    # Append .mat to the segmentation filename
    segmentation_fname_with_extension = f"{segmentation_fname}.mat"
    
    # Construct full path to segmentation file in the source directory
    source_file = os.path.join(segmentation_directory, segmentation_fname_with_extension)
    
    # Check if the file exists in the segmentation directory
    if os.path.exists(source_file):
        # Copy the file to the destination directory
        shutil.copy(source_file, dest_directory)
        print(f"Copied {segmentation_fname_with_extension} to {dest_directory}")
    else:
        print(f"Segmentation file {segmentation_fname_with_extension} not found in {segmentation_directory}")

# Function to copy both .adicht and .mat PV files to the "PV Data" folder inside the destination directory
def copy_PV_files(PV_fname, PV_directory, dest_directory):
    # Create a "PV Data" subfolder inside the destination directory
    pv_data_folder = os.path.join(dest_directory, "PV Data")
    os.makedirs(pv_data_folder, exist_ok=True)
    
    # Define the two file extensions we are looking for
    file_extensions = [".adicht", ".mat"]
    
    # Attempt to copy both files
    for ext in file_extensions:
        source_file = os.path.join(PV_directory, PV_fname + ext)
        
        # Check if the file exists in the PV directory
        if os.path.exists(source_file):
            # Copy the file to the "PV Data" subfolder
            shutil.copy(source_file, pv_data_folder)
            print(f"Copied {PV_fname + ext} to {pv_data_folder}")
        else:
            print(f"PV file {PV_fname + ext} not found in {PV_directory}")

# Updated function to handle directory path generation and optional directory creation
def organise_folders(sheet, outdir, sample_name, mkdir_flag, segmentation_directory, PV_directory):
    if sample_name.lower() == "all":
        # Process all rows if sample_name is "all"
        for row in sheet.iter_rows(min_row=2, values_only=False):
            name, category, weeks, segmentation_fname = get_name_category_week_segmentation(row)
            PV_data = row[1].value  # PV data is the second column
            PV_fname = get_PV_fname(PV_data)
            
            output_directory = get_output_directory(outdir, name, category, weeks)
            if output_directory:
                print(f"{name}: {output_directory}")
                if mkdir_flag:
                    os.makedirs(output_directory, exist_ok=True)
                    # Copy the segmentation file to the created directory
                    copy_segmentation_file(segmentation_fname, segmentation_directory, output_directory)
                    # Copy the PV files (.adicht and .mat) to the "PV Data" subfolder
                    if PV_fname:
                        copy_PV_files(PV_fname, PV_directory, output_directory)
    else:
        # Process only the specific sample
        for row in sheet.iter_rows(min_row=2, values_only=False):
            name, category, weeks, segmentation_fname = get_name_category_week_segmentation(row)
            PV_data = row[1].value  # PV data is the second column
            PV_fname = get_PV_fname(PV_data)
            
            # Check if the current row matches the sample name
            if name == sample_name:
                output_directory = get_output_directory(outdir, name, category, weeks)
                if output_directory:
                    print(output_directory)
                    if mkdir_flag:
                        os.makedirs(output_directory, exist_ok=True)
                        # Copy the segmentation file to the created directory
                        copy_segmentation_file(segmentation_fname, segmentation_directory, output_directory)
                        # Copy the PV files (.adicht and .mat) to the "PV Data" subfolder
                        if PV_fname:
                            copy_PV_files(PV_fname, PV_directory, output_directory)
                return
        
        print(f"Sample {sample_name} not found.")

def main(args=None) -> int:
    """
    Parse the command-line arguments.
    """
    parser = argparse.ArgumentParser()    
    parser.add_argument(
        "-e",
        "--excel",
        default="/home/shared/dynacomp/00_data/CineData/MasterSheet_Code.xlsx",
        type=str,
        help="The full address to the excel file",
    )
    
    parser.add_argument(
        "-s",
        "--sample",
        default="OP100.1",
        type=str,
        help="The sample name or 'all' to process all samples",
    )
    
    parser.add_argument(
        "-o",
        "--outdir",
        default="/home/shared/dynacomp/00_data/CineData/",
        type=str,
        help="The full address to the output directory",
    )
    
    parser.add_argument(
        "-m",
        "--mkdir",
        action='store_true',
        help="Flag to create directories if set",
    )
    
    parser.add_argument(
        "-sd",
        "--segdir",
        default="/home/shared/dynacomp/00_data/CineData/Raw Data/Segmentation",
        type=str,
        help="The directory where segmentation files are located",
    )
    
    parser.add_argument(
        "-pd",
        "--pvdir",
        default="/home/shared/dynacomp/00_data/CineData/Raw Data/PV",
        type=str,
        help="The directory where PV files are located",
    )
    
    args = parser.parse_args(args)
    excel_file_path = args.excel
    sample_name = args.sample
    outdir = args.outdir
    mkdir_flag = args.mkdir
    segmentation_directory = args.segdir
    PV_directory = args.pvdir
    
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    
    # Call the organise_folders function to handle the directory path generation, directory creation, and file copying
    organise_folders(sheet, outdir, sample_name, mkdir_flag, segmentation_directory, PV_directory)

if __name__ == "__main__":
    main()
